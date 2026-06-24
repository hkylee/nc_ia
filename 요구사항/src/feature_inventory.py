"""Feature inventory workbook cleaner and SQLite indexer."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Optional, Sequence

from openpyxl import load_workbook

try:
    from runtime_paths import FEATURE_INVENTORY_DB_PATH, INPUT_ROOT
except ImportError:  # pragma: no cover - package import fallback.
    from .runtime_paths import FEATURE_INVENTORY_DB_PATH, INPUT_ROOT


FEATURE_INVENTORY_DB_SCHEMA_VERSION = 1
DEFAULT_FEATURE_INVENTORY_WORKBOOK = INPUT_ROOT / "references" / "SKT_T4S_기능내역서_정리.xlsx"
DEFAULT_FEATURE_INVENTORY_DB_PATH = FEATURE_INVENTORY_DB_PATH
EXCEL_ERROR_VALUES = {"#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!"}
HEADER_ALIASES = {
    "screen_id": ("screen_id_new", "screen_id", "화면id", "화면 id"),
    "screen_name": ("screen_name", "화면명"),
    "feature_name": ("기능명", "feature_name"),
    "feature_description": ("기능 세부 설명", "기능세부설명", "feature_description"),
    "depth1": ("1Depth", "1depth", "depth1"),
    "depth2": ("2Depth", "2depth", "depth2"),
    "depth3": ("3Depth", "3depth", "depth3"),
    "depth4": ("4Depth", "4depth", "depth4"),
    "depth5": ("5Depth", "5depth", "depth5"),
    "status": ("진행 상태", "진행상태", "status"),
    "scroll": ("scroll", "스크롤"),
    "confidence_level": ("confidence level", "confidence_level", "신뢰도"),
    "confidence_basis": ("confidence basis", "confidence_basis", "신뢰도 근거"),
}
REQUIRED_HEADER_KEYS = {"feature_name", "feature_description"}


def ensure_feature_inventory_database(
    workbook_path: Path = DEFAULT_FEATURE_INVENTORY_WORKBOOK,
    database_path: Path = DEFAULT_FEATURE_INVENTORY_DB_PATH,
    *,
    force: bool = False,
) -> Path:
    workbook_path = Path(workbook_path)
    database_path = Path(database_path)
    database_path.parent.mkdir(parents=True, exist_ok=True)
    source_hash = file_sha256(workbook_path)
    with sqlite3.connect(database_path) as conn:
        conn.row_factory = sqlite3.Row
        initialize_feature_inventory_database(conn)
        if not force and feature_inventory_database_is_current(conn, workbook_path, source_hash):
            return database_path
        rebuild_feature_inventory_database(conn, workbook_path, source_hash)
    return database_path


def initialize_feature_inventory_database(conn: sqlite3.Connection) -> None:
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS metadata (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS feature_rows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            source_sheet TEXT NOT NULL,
            source_row_number INTEGER NOT NULL,
            channel TEXT NOT NULL,
            screen_id TEXT NOT NULL,
            screen_name TEXT NOT NULL,
            feature_name TEXT NOT NULL,
            feature_description TEXT NOT NULL,
            condition_text TEXT NOT NULL,
            input_text TEXT NOT NULL,
            output_text TEXT NOT NULL,
            depth1 TEXT NOT NULL,
            depth2 TEXT NOT NULL,
            depth3 TEXT NOT NULL,
            depth4 TEXT NOT NULL,
            depth5 TEXT NOT NULL,
            depth_path TEXT NOT NULL,
            normalized_depth_path TEXT NOT NULL,
            status TEXT NOT NULL,
            scroll TEXT NOT NULL,
            scroll_value REAL,
            confidence_level TEXT NOT NULL,
            confidence_basis TEXT NOT NULL,
            confidence_score INTEGER,
            canonical_key TEXT NOT NULL,
            row_hash TEXT NOT NULL,
            duplicate_index INTEGER NOT NULL,
            is_duplicate INTEGER NOT NULL,
            search_text TEXT NOT NULL,
            raw_json TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS cleanup_issues (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            source_sheet TEXT NOT NULL,
            source_row_number INTEGER NOT NULL,
            issue_type TEXT NOT NULL,
            message TEXT NOT NULL,
            raw_json TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS feature_channel_summary (
            channel TEXT PRIMARY KEY,
            source_rows INTEGER NOT NULL,
            feature_rows INTEGER NOT NULL,
            unique_feature_rows INTEGER NOT NULL,
            duplicate_rows INTEGER NOT NULL,
            screen_count INTEGER NOT NULL,
            depth1_count INTEGER NOT NULL,
            issue_count INTEGER NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS feature_screen_summary (
            channel TEXT NOT NULL,
            screen_id TEXT NOT NULL,
            screen_name TEXT NOT NULL,
            feature_rows INTEGER NOT NULL,
            unique_feature_rows INTEGER NOT NULL,
            PRIMARY KEY (channel, screen_id, screen_name)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS feature_depth_summary (
            channel TEXT NOT NULL,
            depth_level INTEGER NOT NULL,
            depth_value TEXT NOT NULL,
            feature_rows INTEGER NOT NULL,
            unique_feature_rows INTEGER NOT NULL,
            PRIMARY KEY (channel, depth_level, depth_value)
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_feature_rows_channel ON feature_rows(channel)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_feature_rows_screen ON feature_rows(screen_id, screen_name)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_feature_rows_depth_path ON feature_rows(normalized_depth_path)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_feature_rows_feature_name ON feature_rows(feature_name)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_feature_rows_canonical ON feature_rows(canonical_key)")
    conn.execute("DROP VIEW IF EXISTS feature_unique_rows")
    conn.execute("CREATE VIEW feature_unique_rows AS SELECT * FROM feature_rows WHERE is_duplicate = 0")
    conn.commit()


def feature_inventory_database_is_current(conn: sqlite3.Connection, workbook_path: Path, source_hash: str) -> bool:
    metadata = dict(conn.execute("SELECT key, value FROM metadata").fetchall())
    return (
        metadata.get("schema_version") == str(FEATURE_INVENTORY_DB_SCHEMA_VERSION)
        and metadata.get("source_path") == str(workbook_path.resolve(strict=False))
        and metadata.get("source_hash") == source_hash
    )


def rebuild_feature_inventory_database(conn: sqlite3.Connection, workbook_path: Path, source_hash: str) -> None:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    conn.execute("DELETE FROM feature_rows")
    conn.execute("DELETE FROM cleanup_issues")
    conn.execute("DELETE FROM feature_channel_summary")
    conn.execute("DELETE FROM feature_screen_summary")
    conn.execute("DELETE FROM feature_depth_summary")
    conn.execute("DELETE FROM metadata")

    seen_keys: Counter[str] = Counter()
    inserted_rows: list[dict[str, Any]] = []
    issue_rows: list[dict[str, Any]] = []
    source_row_counts: Counter[str] = Counter()

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = list(iter_worksheet_rows(worksheet))
        if not rows:
            continue
        header_row_index, headers = detect_header_row(rows)
        if header_row_index is None:
            issue_rows.append(cleanup_issue(workbook_path, sheet_name, 0, "header_missing", "기능명/기능 세부 설명 헤더를 찾지 못했습니다.", {}))
            continue
        indexes = feature_column_indexes(headers)
        source_rows = rows[header_row_index + 1 :]
        source_row_counts[sheet_name] += len(source_rows)
        for offset, raw_row in enumerate(source_rows, start=header_row_index + 2):
            raw_json = raw_row_json(headers, raw_row)
            cleaned_row, errors = clean_row_values(raw_row)
            if errors:
                issue_rows.append(
                    cleanup_issue(
                        workbook_path,
                        sheet_name,
                        offset,
                        "excel_error_cleaned",
                        f"{', '.join(sorted(errors))} 값을 빈 값으로 정제했습니다.",
                        raw_json,
                    )
                )
            if not any(clean_text(value) for value in cleaned_row):
                continue
            record = feature_record_from_row(workbook_path, sheet_name, offset, cleaned_row, raw_json, indexes)
            if not record["feature_name"] and not record["feature_description"]:
                continue
            if not record["feature_name"]:
                issue_rows.append(cleanup_issue(workbook_path, sheet_name, offset, "missing_feature_name", "기능명이 비어 있습니다.", raw_json))
                continue
            if not record["feature_description"]:
                issue_rows.append(
                    cleanup_issue(workbook_path, sheet_name, offset, "missing_feature_description", "기능 세부 설명이 비어 있습니다.", raw_json)
                )
            seen_keys[record["canonical_key"]] += 1
            record["duplicate_index"] = seen_keys[record["canonical_key"]]
            record["is_duplicate"] = 1 if record["duplicate_index"] > 1 else 0
            if record["is_duplicate"]:
                issue_rows.append(cleanup_issue(workbook_path, sheet_name, offset, "duplicate_canonical_key", "동일 기능 행으로 판정되어 중복 표시했습니다.", raw_json))
            inserted_rows.append(record)
            conn.execute(
                """
                INSERT INTO feature_rows (
                    source_file, source_sheet, source_row_number, channel, screen_id, screen_name,
                    feature_name, feature_description, condition_text, input_text, output_text,
                    depth1, depth2, depth3, depth4, depth5, depth_path, normalized_depth_path,
                    status, scroll, scroll_value, confidence_level, confidence_basis, confidence_score,
                    canonical_key, row_hash, duplicate_index, is_duplicate, search_text, raw_json
                )
                VALUES (
                    :source_file, :source_sheet, :source_row_number, :channel, :screen_id, :screen_name,
                    :feature_name, :feature_description, :condition_text, :input_text, :output_text,
                    :depth1, :depth2, :depth3, :depth4, :depth5, :depth_path, :normalized_depth_path,
                    :status, :scroll, :scroll_value, :confidence_level, :confidence_basis, :confidence_score,
                    :canonical_key, :row_hash, :duplicate_index, :is_duplicate, :search_text, :raw_json
                )
                """,
                record,
            )

    for issue in issue_rows:
        conn.execute(
            """
            INSERT INTO cleanup_issues (
                source_file, source_sheet, source_row_number, issue_type, message, raw_json
            )
            VALUES (
                :source_file, :source_sheet, :source_row_number, :issue_type, :message, :raw_json
            )
            """,
            issue,
        )

    write_summary_tables(conn, source_row_counts, inserted_rows, issue_rows)
    write_feature_inventory_metadata(conn, workbook_path, source_hash, workbook.sheetnames, source_row_counts, inserted_rows, issue_rows)
    conn.commit()


def iter_worksheet_rows(worksheet: Any) -> Iterable[list[Any]]:
    for row in worksheet.iter_rows(values_only=True):
        yield list(row)


def detect_header_row(rows: Sequence[Sequence[Any]]) -> tuple[Optional[int], list[str]]:
    for index, row in enumerate(rows[:20]):
        headers = make_unique_headers([clean_text(value) for value in row])
        keys = set(feature_column_indexes(headers).values())
        mapped = set(feature_column_indexes(headers).keys())
        if REQUIRED_HEADER_KEYS.issubset(mapped) and keys:
            return index, headers
    return None, []


def feature_column_indexes(headers: Sequence[str]) -> dict[str, int]:
    normalized = {normalize_header(header): index for index, header in enumerate(headers)}
    result: dict[str, int] = {}
    for key, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            index = normalized.get(normalize_header(alias))
            if index is not None:
                result[key] = index
                break
    return result


def feature_record_from_row(
    workbook_path: Path,
    sheet_name: str,
    row_number: int,
    row: Sequence[str],
    raw_json: Mapping[str, str],
    indexes: Mapping[str, int],
) -> dict[str, Any]:
    channel = channel_from_sheet(sheet_name)
    screen_id = cell_text(row, indexes.get("screen_id", -1))
    screen_name = cell_text(row, indexes.get("screen_name", -1))
    feature_name = cell_text(row, indexes.get("feature_name", -1))
    feature_description = normalize_description(cell_text(row, indexes.get("feature_description", -1)))
    parsed = parse_feature_description(feature_description)
    depths = [cell_text(row, indexes.get(f"depth{level}", -1)) for level in range(1, 6)]
    depth_path = " > ".join([value for value in depths if value])
    status = cell_text(row, indexes.get("status", -1))
    scroll = cell_text(row, indexes.get("scroll", -1))
    confidence_level = cell_text(row, indexes.get("confidence_level", -1))
    confidence_basis = normalize_description(cell_text(row, indexes.get("confidence_basis", -1)))
    confidence_score = extract_confidence_score(confidence_basis)
    significant = [
        channel,
        screen_id,
        screen_name,
        feature_name,
        feature_description,
        *depths,
    ]
    canonical_key = sha1_text("|".join(normalize_match(value) for value in significant))
    row_hash = sha1_text(json.dumps([channel, screen_id, screen_name, feature_name, feature_description, *depths], ensure_ascii=False))
    search_text = clean_text(" ".join([channel, screen_id, screen_name, feature_name, feature_description, depth_path, status]))
    return {
        "source_file": workbook_path.name,
        "source_sheet": sheet_name,
        "source_row_number": row_number,
        "channel": channel,
        "screen_id": screen_id,
        "screen_name": screen_name,
        "feature_name": feature_name,
        "feature_description": feature_description,
        "condition_text": parsed.get("조건", ""),
        "input_text": parsed.get("입력", ""),
        "output_text": parsed.get("출력", ""),
        "depth1": depths[0],
        "depth2": depths[1],
        "depth3": depths[2],
        "depth4": depths[3],
        "depth5": depths[4],
        "depth_path": depth_path,
        "normalized_depth_path": normalize_match(depth_path),
        "status": status,
        "scroll": scroll,
        "scroll_value": parse_float(scroll),
        "confidence_level": confidence_level,
        "confidence_basis": confidence_basis,
        "confidence_score": confidence_score,
        "canonical_key": canonical_key,
        "row_hash": row_hash,
        "duplicate_index": 1,
        "is_duplicate": 0,
        "search_text": search_text,
        "raw_json": json.dumps(raw_json, ensure_ascii=False, sort_keys=True),
    }


def write_summary_tables(
    conn: sqlite3.Connection,
    source_row_counts: Mapping[str, int],
    records: Sequence[Mapping[str, Any]],
    issues: Sequence[Mapping[str, Any]],
) -> None:
    by_channel: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    issue_counts: Counter[str] = Counter()
    for record in records:
        by_channel[str(record["channel"])].append(record)
    for issue in issues:
        issue_counts[channel_from_sheet(str(issue["source_sheet"]))] += 1
    for channel, rows in sorted(by_channel.items()):
        screen_keys = {
            (str(row["screen_id"]), str(row["screen_name"]))
            for row in rows
            if row.get("screen_id") or row.get("screen_name")
        }
        depth1_values = {str(row["depth1"]) for row in rows if row.get("depth1")}
        duplicate_rows = sum(1 for row in rows if int(row["is_duplicate"]) == 1)
        conn.execute(
            """
            INSERT INTO feature_channel_summary (
                channel, source_rows, feature_rows, unique_feature_rows, duplicate_rows, screen_count, depth1_count, issue_count
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                channel,
                source_row_counts.get(channel, source_row_counts.get(sheet_from_channel(channel), 0)),
                len(rows),
                len(rows) - duplicate_rows,
                duplicate_rows,
                len(screen_keys),
                len(depth1_values),
                issue_counts.get(channel, 0),
            ),
        )

    screen_counts: Counter[tuple[str, str, str]] = Counter()
    screen_unique_counts: Counter[tuple[str, str, str]] = Counter()
    depth_counts: Counter[tuple[str, int, str]] = Counter()
    depth_unique_counts: Counter[tuple[str, int, str]] = Counter()
    for row in records:
        channel = str(row["channel"])
        screen_key = (channel, str(row["screen_id"]), str(row["screen_name"]))
        screen_counts[screen_key] += 1
        if not int(row["is_duplicate"]):
            screen_unique_counts[screen_key] += 1
        for level in range(1, 6):
            value = str(row.get(f"depth{level}", "") or "")
            if not value:
                continue
            depth_key = (channel, level, value)
            depth_counts[depth_key] += 1
            if not int(row["is_duplicate"]):
                depth_unique_counts[depth_key] += 1
    for key, count in sorted(screen_counts.items()):
        conn.execute(
            "INSERT INTO feature_screen_summary (channel, screen_id, screen_name, feature_rows, unique_feature_rows) VALUES (?, ?, ?, ?, ?)",
            (*key, count, screen_unique_counts.get(key, 0)),
        )
    for key, count in sorted(depth_counts.items()):
        conn.execute(
            "INSERT INTO feature_depth_summary (channel, depth_level, depth_value, feature_rows, unique_feature_rows) VALUES (?, ?, ?, ?, ?)",
            (*key, count, depth_unique_counts.get(key, 0)),
        )


def write_feature_inventory_metadata(
    conn: sqlite3.Connection,
    workbook_path: Path,
    source_hash: str,
    sheet_names: Sequence[str],
    source_row_counts: Mapping[str, int],
    records: Sequence[Mapping[str, Any]],
    issues: Sequence[Mapping[str, Any]],
) -> None:
    duplicate_count = sum(1 for row in records if int(row["is_duplicate"]) == 1)
    metadata = {
        "schema_version": str(FEATURE_INVENTORY_DB_SCHEMA_VERSION),
        "source_path": str(workbook_path.resolve(strict=False)),
        "source_name": workbook_path.name,
        "source_hash": source_hash,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "sheet_names": json.dumps(list(sheet_names), ensure_ascii=False),
        "source_row_count": str(sum(source_row_counts.values())),
        "feature_row_count": str(len(records)),
        "unique_feature_row_count": str(len(records) - duplicate_count),
        "duplicate_row_count": str(duplicate_count),
        "cleanup_issue_count": str(len(issues)),
    }
    conn.executemany("INSERT OR REPLACE INTO metadata(key, value) VALUES (?, ?)", sorted(metadata.items()))


def inspect_feature_inventory_database(database_path: Path = DEFAULT_FEATURE_INVENTORY_DB_PATH) -> dict[str, Any]:
    if not Path(database_path).exists():
        return {"ok": False, "error": "feature_inventory.db not found", "path": str(database_path)}
    with sqlite3.connect(database_path) as conn:
        conn.row_factory = sqlite3.Row
        metadata = dict(conn.execute("SELECT key, value FROM metadata").fetchall())
        channels = [dict(row) for row in conn.execute("SELECT * FROM feature_channel_summary ORDER BY channel").fetchall()]
        issue_types = [
            dict(row)
            for row in conn.execute(
                "SELECT issue_type, COUNT(*) AS count FROM cleanup_issues GROUP BY issue_type ORDER BY count DESC, issue_type"
            ).fetchall()
        ]
    return {"ok": True, "path": str(database_path), "metadata": metadata, "channels": channels, "issueTypes": issue_types}


def cleanup_issue(
    workbook_path: Path,
    sheet_name: str,
    row_number: int,
    issue_type: str,
    message: str,
    raw_json: Mapping[str, Any],
) -> dict[str, Any]:
    return {
        "source_file": workbook_path.name,
        "source_sheet": sheet_name,
        "source_row_number": row_number,
        "issue_type": issue_type,
        "message": message,
        "raw_json": json.dumps(raw_json, ensure_ascii=False, sort_keys=True),
    }


def raw_row_json(headers: Sequence[str], row: Sequence[Any]) -> dict[str, str]:
    result: dict[str, str] = {}
    for index, header in enumerate(headers):
        value = clean_cell_value(row[index] if index < len(row) else None)
        if header and value:
            result[header] = value
    return result


def clean_row_values(row: Sequence[Any]) -> tuple[list[str], set[str]]:
    errors: set[str] = set()
    cleaned: list[str] = []
    for value in row:
        text = clean_cell_value(value)
        if text in EXCEL_ERROR_VALUES:
            errors.add(text)
            text = ""
        cleaned.append(text)
    return cleaned, errors


def clean_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = unicodedata.normalize("NFC", str(value))
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\u00a0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r" *\n *", "\n", text)
    return text.strip()


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_cell_value(value)).strip()


def normalize_description(value: str) -> str:
    text = clean_cell_value(value)
    lines = [clean_text(line) for line in text.split("\n")]
    return "\n".join(line for line in lines if line)


def parse_feature_description(description: str) -> dict[str, str]:
    result = {"조건": "", "입력": "", "출력": ""}
    if not description:
        return result
    pattern = re.compile(r"\[(조건|입력|출력)\]\s*")
    matches = list(pattern.finditer(description))
    if not matches:
        return result
    for index, match in enumerate(matches):
        key = match.group(1)
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(description)
        value = clean_text(description[start:end])
        if not value:
            continue
        result[key] = f"{result[key]} {value}".strip() if result[key] else value
    return result


def cell_text(row: Sequence[str], index: int) -> str:
    if index < 0 or index >= len(row):
        return ""
    return clean_cell_value(row[index])


def make_unique_headers(headers: Sequence[str]) -> list[str]:
    counts: Counter[str] = Counter()
    result: list[str] = []
    for index, header in enumerate(headers):
        base = clean_text(header) or f"column_{index + 1}"
        counts[base] += 1
        result.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return result


def normalize_header(value: str) -> str:
    return re.sub(r"[\s_]+", "", clean_text(value).lower())


def normalize_match(value: Any) -> str:
    text = unicodedata.normalize("NFC", clean_text(value)).lower()
    return re.sub(r"[^0-9a-z가-힣]+", "", text)


def channel_from_sheet(sheet_name: str) -> str:
    return "통합" if sheet_name == "통합 기능 목록" else sheet_name


def sheet_from_channel(channel: str) -> str:
    return "통합 기능 목록" if channel == "통합" else channel


def parse_float(value: str) -> Optional[float]:
    text = clean_text(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def extract_confidence_score(value: str) -> Optional[int]:
    match = re.search(r"\[(\d{1,3})%\]", value)
    if match:
        return max(0, min(100, int(match.group(1))))
    return None


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha1_text(value: str) -> str:
    return hashlib.sha1(value.encode("utf-8")).hexdigest()


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build cleaned feature inventory SQLite database.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_FEATURE_INVENTORY_WORKBOOK)
    parser.add_argument("--database", type=Path, default=DEFAULT_FEATURE_INVENTORY_DB_PATH)
    parser.add_argument("--force", action="store_true")
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = build_arg_parser().parse_args(argv)
    database_path = ensure_feature_inventory_database(args.workbook, args.database, force=args.force)
    summary = inspect_feature_inventory_database(database_path)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
