import sqlite3

from openpyxl import Workbook

from src.feature_inventory import ensure_feature_inventory_database, inspect_feature_inventory_database


def test_feature_inventory_database_cleans_and_indexes_workbook(tmp_path):
    workbook_path = tmp_path / "features.xlsx"
    db_path = tmp_path / "feature_inventory.db"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "T 월드"
    sheet.append(["", "", "", ""])
    sheet.append(["screen_id_new", "screen_name", "기능명", "기능 세부 설명", "1Depth", "2Depth", "진행 상태"])
    sheet.append(["#REF!", "", "", "", "", "", ""])
    sheet.append(
        [
            "TW-001",
            "메인",
            "검색 화면 이동",
            "[조건] 로그인 상태\n[입력] 검색 아이콘 선택\n[출력] 검색 화면으로 이동",
            "메인",
            "헤더",
            "리뷰중",
        ]
    )
    sheet.append(
        [
            "TW-001",
            "메인",
            "검색 화면 이동",
            "[조건] 로그인 상태\n[입력] 검색 아이콘 선택\n[출력] 검색 화면으로 이동",
            "메인",
            "헤더",
            "리뷰중",
        ]
    )
    integrated = workbook.create_sheet("통합 기능 목록")
    integrated.append(["기능명", "기능 세부 설명", "scroll", "confidence level", "confidence basis"])
    integrated.append(["소셜 로그인 수단 노출", "[출력] 로그인 수단 목록 노출", "0.05", "C0", "[90%]\n[D] 확인"])
    workbook.save(workbook_path)

    result_path = ensure_feature_inventory_database(workbook_path, db_path, force=True)

    assert result_path == db_path
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT channel, screen_id, feature_name, condition_text, input_text, output_text,
                   depth_path, confidence_score, is_duplicate
            FROM feature_rows
            ORDER BY id
            """
        ).fetchall()
        issues = conn.execute("SELECT issue_type, COUNT(*) AS count FROM cleanup_issues GROUP BY issue_type").fetchall()
        metadata = dict(conn.execute("SELECT key, value FROM metadata").fetchall())
        unique_count = conn.execute("SELECT COUNT(*) FROM feature_unique_rows").fetchone()[0]

    assert len(rows) == 3
    assert rows[0]["channel"] == "T 월드"
    assert rows[0]["screen_id"] == "TW-001"
    assert rows[0]["condition_text"] == "로그인 상태"
    assert rows[0]["input_text"] == "검색 아이콘 선택"
    assert rows[0]["output_text"] == "검색 화면으로 이동"
    assert rows[0]["depth_path"] == "메인 > 헤더"
    assert rows[1]["is_duplicate"] == 1
    assert rows[2]["channel"] == "통합"
    assert rows[2]["confidence_score"] == 90
    assert unique_count == 2
    assert metadata["schema_version"] == "1"
    assert {issue["issue_type"]: issue["count"] for issue in issues}["excel_error_cleaned"] == 1
    assert {issue["issue_type"]: issue["count"] for issue in issues}["duplicate_canonical_key"] == 1


def test_feature_inventory_inspect_returns_summary(tmp_path):
    workbook_path = tmp_path / "features.xlsx"
    db_path = tmp_path / "feature_inventory.db"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "T 다이렉트"
    sheet.append(["screen_id_new", "screen_name", "기능명", "기능 세부 설명", "1Depth"])
    sheet.append(["TD-001", "메인", "전체 메뉴 이동", "[입력] 메뉴 선택\n[출력] 전체 메뉴 화면으로 이동", "메인"])
    workbook.save(workbook_path)

    ensure_feature_inventory_database(workbook_path, db_path, force=True)
    summary = inspect_feature_inventory_database(db_path)

    assert summary["ok"] is True
    assert summary["metadata"]["feature_row_count"] == "1"
    assert summary["channels"][0]["channel"] == "T 다이렉트"
    assert summary["channels"][0]["feature_rows"] == 1
